from pathlib import Path

from openpyxl import load_workbook

from rsac_relatorios_risco.config.models import ConfigItem, ConfigWorkbook
from rsac_relatorios_risco.config.placeholder_resolver import resolve_value


def load_config_workbook(path: Path, mes: str, ano: str) -> ConfigWorkbook:
    workbook = load_workbook(path, data_only=False)
    context = {"Data": f"{mes}{ano}", "YYYY-MM": f"{ano}-{mes}"}
    settings = _load_settings(workbook)
    sheet = _select_items_sheet(workbook)

    items = []
    for row in _iter_item_rows(sheet):
        merged_row = _merge_row_with_settings(row, settings)
        resolved_row = _resolve_row(merged_row, context)
        if not resolved_row["Reference"].strip():
            raise ValueError("Reference vazia após resolução")
        items.append(_build_item(resolved_row))

    resolved_settings = {
        key: resolve_value(value, context)
        for key, value in settings.items()
    }
    return ConfigWorkbook(settings=resolved_settings, items=items)


def _load_settings(workbook) -> dict[str, str]:
    if "Settings" not in workbook.sheetnames:
        return {}

    sheet = workbook["Settings"]
    settings: dict[str, str] = {}
    for name, value, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if name is None or str(name).strip() == "":
            continue
        settings[str(name)] = "" if value is None else str(value)
    return settings


def _select_items_sheet(workbook):
    if "Items" in workbook.sheetnames:
        return workbook["Items"]
    if "QueueItems" in workbook.sheetnames:
        return workbook["QueueItems"]
    raise ValueError("Nenhuma aba Items ou QueueItems encontrada")


def _iter_item_rows(sheet):
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values, strict=False))
        if any(not _is_empty(value) for value in row.values()):
            yield row


def _merge_row_with_settings(
    row: dict[str, object | None],
    settings: dict[str, str],
) -> dict[str, object | None]:
    merged = dict(row)
    for key, value in row.items():
        if _is_empty(value) and key in settings:
            merged[key] = settings[key]
    return merged


def _resolve_row(
    row: dict[str, object | None],
    context: dict[str, str],
) -> dict[str, str]:
    resolved: dict[str, str] = {}
    for key, value in row.items():
        resolved[key] = resolve_value(value, context) if isinstance(value, str) else (
            "" if value is None else str(value)
        )
    return resolved


def _build_item(row: dict[str, str]) -> ConfigItem:
    return ConfigItem(
        reference=row["Reference"],
        tipo_relatorio=_optional_string(row.get("Tipo Relatorio")),
        timeout=_optional_int(row.get("Timeout")),
        cooperativa=_optional_string(row.get("Cooperativa")),
        pa=_optional_string(row.get("PA")),
        nome_cooperativa_1=_optional_string(row.get("Nome Cooperativa 1")),
        nome_cooperativa_2=_optional_string(row.get("Nome Cooperativa 2")),
        destinatarios=_optional_string(row.get("Destinatarios")),
        sharepoint=_optional_string(row.get("Sharepoint")),
        nome_arquivo=_optional_string(row.get("Nome Arquivo")),
        extensao=_optional_string(row.get("Extensao")),
    )


def _optional_string(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def _optional_int(value: str | None) -> int | None:
    if value is None:
        return None
    stripped = value.strip()
    return int(stripped) if stripped else None


def _is_empty(value: object | None) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")
