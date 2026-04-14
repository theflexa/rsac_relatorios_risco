from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook


HEADER_ROW_INDEX = 6
DATA_START_ROW_INDEX = 7
WORKSHEET_NAME = "Relatório Database"


@dataclass(slots=True)
class ReportData:
    cooperativa: str
    competencia: str
    headers: list[str]
    rows: list[list[str]]
    data_emissao: str
    criterios: str
    all_rows: list[list]


def read_report(path: Path) -> ReportData:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[WORKSHEET_NAME]
    headers = _read_headers(sheet)
    rows = _read_rows(sheet, len(headers))
    cooperativa = _extract_cooperativa(rows, sheet["B5"].value)
    competencia = _extract_competencia(rows)
    data_emissao = str(sheet["B4"].value or "")
    criterios = str(sheet["B5"].value or "")
    all_rows = _read_all_rows(sheet)
    return ReportData(
        cooperativa=cooperativa,
        competencia=competencia,
        headers=headers,
        rows=rows,
        data_emissao=data_emissao,
        criterios=criterios,
        all_rows=all_rows,
    )


def _read_headers(sheet) -> list[str]:
    headers: list[str] = []
    column_index = 1
    while True:
        value = sheet.cell(row=HEADER_ROW_INDEX, column=column_index).value
        if value is None:
            break
        headers.append(str(value))
        column_index += 1
    return headers


def _read_rows(sheet, column_count: int) -> list[list[str]]:
    rows: list[list[str]] = []
    row_index = DATA_START_ROW_INDEX
    while True:
        values = [
            sheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, column_count + 1)
        ]
        if all(value is None for value in values):
            break
        rows.append(["" if value is None else str(value) for value in values])
        row_index += 1
    return rows


def _read_all_rows(sheet) -> list[list]:
    """Le todas as linhas da aba (row 1 ate a ultima com dados), preservando valores originais."""
    max_col = sheet.max_column or 1
    result: list[list] = []
    for row_index in range(1, sheet.max_row + 1):
        values = [sheet.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in values):
            break
        result.append(values)
    return result


def _extract_cooperativa(rows: list[list[str]], criteria_value: object | None) -> str:
    primary_value = rows[0][1] if rows else ""
    primary_match = re.match(r"(?P<codigo>\d+)\s+-", primary_value)
    criteria_match = re.search(r"SINGULAR:\s*(?P<codigo>\d+)", str(criteria_value or ""))

    if primary_match and criteria_match and primary_match.group("codigo") != criteria_match.group("codigo"):
        raise ValueError("Cooperativa do relatório diverge do critério auxiliar")

    if primary_match:
        return primary_match.group("codigo")
    if criteria_match:
        return criteria_match.group("codigo")
    raise ValueError("Não foi possível identificar a cooperativa no relatório")


def _extract_competencia(rows: list[list[str]]) -> str:
    if not rows or len(rows[0]) < 5 or not rows[0][4].strip():
        raise ValueError("Não foi possível identificar a competência no relatório")
    return rows[0][4]
