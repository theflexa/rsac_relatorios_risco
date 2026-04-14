from pathlib import Path

from openpyxl import load_workbook

from rsac_relatorios_risco.services.report_service import ReportData


class ConsolidadoService:
    def apply_report(self, workbook_path: Path, report: ReportData) -> None:
        """Copia todo o conteudo do insumo para a aba da cooperativa no consolidado."""
        workbook = load_workbook(workbook_path)
        sheet = workbook[report.cooperativa]

        # Desfazer merges antes de limpar
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))

        # Limpar conteudo existente na aba
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        # Copiar todas as linhas do insumo
        for row_index, row_values in enumerate(report.all_rows, start=1):
            for col_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=col_index).value = value

        workbook.save(workbook_path)


# Wrapper de modulo para uso direto: apply_report(path, report)
_default_service = ConsolidadoService()


def apply_report(workbook_path: Path, report: ReportData) -> None:
    _default_service.apply_report(workbook_path, report)
