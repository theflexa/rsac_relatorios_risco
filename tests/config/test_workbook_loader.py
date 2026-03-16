from pathlib import Path

import pytest
from openpyxl import Workbook

from rsac_relatorios_risco.config.workbook_loader import load_config_workbook


ITEM_HEADERS = [
    "Reference",
    "Tipo Relatorio",
    "Timeout",
    "Cooperativa",
    "PA",
    "Nome Cooperativa 1",
    "Nome Cooperativa 2",
    "Destinatarios",
    "Sharepoint",
    "Nome Arquivo",
    "Extensao",
]


def _write_workbook(
    tmp_path: Path,
    *,
    settings_rows: list[tuple[str, str | int | None]] | None = None,
    items_rows: list[list[object | None]] | None = None,
    queue_items_rows: list[list[object | None]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    settings_sheet = workbook.create_sheet("Settings")
    settings_sheet.append(["Name", "Value"])
    for row in settings_rows or []:
        settings_sheet.append(row)

    if items_rows is not None:
        items_sheet = workbook.create_sheet("Items")
        items_sheet.append(ITEM_HEADERS)
        for row in items_rows:
            items_sheet.append(row)

    if queue_items_rows is not None:
        queue_sheet = workbook.create_sheet("QueueItems")
        queue_sheet.append(ITEM_HEADERS)
        for row in queue_items_rows:
            queue_sheet.append(row)

    path = tmp_path / "config.xlsx"
    workbook.save(path)
    return path


def test_load_config_workbook_prefers_items_sheet_when_items_and_queueitems_exist(
    tmp_path: Path,
):
    path = _write_workbook(
        tmp_path,
        items_rows=[
            [
                "ITEMS_{Data}",
                "RSAC",
                480,
                "3333",
                "0001",
                "Coop 1",
                "Coop 2",
                "items@sicoob.com",
                "share/{YYYY-MM}",
                '="ITEMS_{Data}"',
                ".xlsx",
            ]
        ],
        queue_items_rows=[
            [
                "QUEUE_{Data}",
                "RSAC",
                480,
                "4444",
                "0002",
                "Fila 1",
                "Fila 2",
                "queue@sicoob.com",
                "queue/{YYYY-MM}",
                '="QUEUE_{Data}"',
                ".xlsx",
            ]
        ],
    )

    workbook = load_config_workbook(path, mes="03", ano="2026")

    assert len(workbook.items) == 1
    assert workbook.items[0].reference == "ITEMS_032026"
    assert workbook.items[0].cooperativa == "3333"


def test_load_config_workbook_falls_back_to_queueitems_when_items_missing(
    tmp_path: Path,
):
    path = _write_workbook(
        tmp_path,
        queue_items_rows=[
            [
                "QUEUE_{Data}",
                "RSAC",
                480,
                "4444",
                "0002",
                "Fila 1",
                "Fila 2",
                "queue@sicoob.com",
                "queue/{YYYY-MM}",
                '="QUEUE_{Data}"',
                ".xlsx",
            ]
        ],
    )

    workbook = load_config_workbook(path, mes="03", ano="2026")

    assert len(workbook.items) == 1
    assert workbook.items[0].reference == "QUEUE_032026"
    assert workbook.items[0].cooperativa == "4444"


def test_load_config_workbook_applies_settings_defaults_to_empty_item_fields(
    tmp_path: Path,
):
    path = _write_workbook(
        tmp_path,
        settings_rows=[
            ("Destinatarios", "default@sicoob.com"),
            ("Sharepoint", "share/{YYYY-MM}"),
            ("Extensao", ".xlsx"),
        ],
        items_rows=[
            [
                '=D2 & "_RSAC_RISCO_{Data}"',
                "RSAC",
                480,
                "3333",
                "0001",
                "Coop 1",
                "Coop 2",
                "",
                None,
                '="ARQUIVO_{YYYY-MM}"',
                "",
            ]
        ],
    )

    workbook = load_config_workbook(path, mes="03", ano="2026")
    item = workbook.items[0]

    assert item.reference == '=D2 & "_RSAC_RISCO_032026"'
    assert item.destinatarios == "default@sicoob.com"
    assert item.sharepoint == "share/2026-03"
    assert item.nome_arquivo == '="ARQUIVO_2026-03"'
    assert item.extensao == ".xlsx"


def test_load_config_workbook_raises_when_reference_is_blank_after_resolution(
    tmp_path: Path,
):
    path = _write_workbook(
        tmp_path,
        items_rows=[
            [
                "",
                "RSAC",
                480,
                "3333",
                "0001",
                "Coop 1",
                "Coop 2",
                "items@sicoob.com",
                "share/{YYYY-MM}",
                '="ITEMS_{Data}"',
                ".xlsx",
            ]
        ],
    )

    with pytest.raises(ValueError, match="Reference vazia"):
        load_config_workbook(path, mes="03", ano="2026")
