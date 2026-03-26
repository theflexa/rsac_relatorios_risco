from pathlib import Path

from openpyxl import Workbook

from rsac_relatorios_risco.dispatcher.service import dispatch_config_items
from rsac_relatorios_risco.integrations.database_client import build_item_payload
from rsac_relatorios_risco.integrations.jarbis_client import build_process_variables


ITEM_HEADERS = [
    "Reference",
    "Tipo Relatorio",
    "Timeout",
    "Cooperativa",
    "PA",
    "Nome Cooperativa 1",
    "Nome Cooperativa 2",
    "Destinatarios",
    "Sharepoint",
    "Nome Arquivo",
    "Extensao",
]


def _write_dispatcher_workbook(
    tmp_path: Path,
    *,
    settings_rows: list[tuple[str, str | int | None]] | None = None,
    items_rows: list[list[object | None]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    settings_sheet = workbook.create_sheet("Settings")
    settings_sheet.append(["Name", "Value"])
    for row in settings_rows or []:
        settings_sheet.append(row)

    items_sheet = workbook.create_sheet("Items")
    items_sheet.append(ITEM_HEADERS)
    for row in items_rows or []:
        items_sheet.append(row)

    path = tmp_path / "config_dispatcher.xlsx"
    workbook.save(path)
    return path


def test_build_item_payload_preserves_sheet_reference():
    payload = build_item_payload(
        1,
        2,
        "3333_RSAC_RISCO_032026",
        {"cooperativa": "3333"},
    )

    assert payload["reference"] == "3333_RSAC_RISCO_032026"
    assert payload["status"] == "pendente"


def test_build_process_variables_uses_inserted_item_reference():
    variables = build_process_variables({"reference": "3333_RSAC_RISCO_032026"})

    assert variables["reference"]["value"] == "3333_RSAC_RISCO_032026"


def test_dispatch_config_items_reuses_existing_reference_and_inserts_only_new_items(
    tmp_path: Path,
):
    path = _write_dispatcher_workbook(
        tmp_path,
        settings_rows=[
            ("Destinatarios", "default@sicoob.com"),
            ("Sharepoint", "share/{YYYY-MM}"),
            ("Extensao", ".xlsx"),
        ],
        items_rows=[
            [
                "3333_RSAC_RISCO_{Data}",
                "RSAC",
                480,
                "3333",
                "0001",
                "Coop 1",
                "Coop 1",
                "",
                "",
                "Arquivo_3333_{Data}",
                "",
            ],
            [
                "4444_RSAC_RISCO_{Data}",
                "RSAC",
                480,
                "4444",
                "0002",
                "Coop 2",
                "Coop 2",
                "",
                "",
                "Arquivo_4444_{Data}",
                "",
            ],
        ],
    )
    inserted_payloads: list[tuple[dict, dict]] = []
    existing_references = {"3333_RSAC_RISCO_032026"}

    result = dispatch_config_items(
        config_path=path,
        mes="03",
        ano="2026",
        project_id=10,
        job_id=20,
        reference_exists=existing_references.__contains__,
        insert_item=lambda payload, variables: inserted_payloads.append(
            (payload, variables),
        ),
    )

    assert result.inserted_count == 1
    assert result.skipped_count == 1
    assert result.reused_references == ["3333_RSAC_RISCO_032026"]
    assert any("reaproveitado" in message for message in result.log_messages)
    assert len(inserted_payloads) == 1
    payload, variables = inserted_payloads[0]
    assert payload["reference"] == "4444_RSAC_RISCO_032026"
    assert payload["status"] == "pendente"
    assert payload["data"]["cooperativa"] == "4444"
    assert payload["data"]["destinatarios"] == "default@sicoob.com"
    assert payload["data"]["sharepoint"] == "share/2026-03"
    assert variables["reference"]["value"] == "4444_RSAC_RISCO_032026"
