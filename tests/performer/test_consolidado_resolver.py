from pathlib import Path

from openpyxl import Workbook

from rsac_relatorios_risco.performer.consolidado_resolver import (
    resolve_monthly_workbook,
)


def test_resolve_monthly_workbook_creates_copy_from_template_when_missing(
    tmp_path: Path,
):
    template = tmp_path / "modelo.xlsx"
    workbook = Workbook()
    workbook.save(template)

    workbook_path = resolve_monthly_workbook(
        template_path=template,
        output_dir=tmp_path,
        competencia="03/2026",
        file_name="RSAC_032026.xlsx",
    )

    assert workbook_path.exists()
    assert workbook_path.name == "RSAC_032026.xlsx"


def test_resolve_monthly_workbook_reuses_existing_file_when_already_created(
    tmp_path: Path,
):
    template = tmp_path / "modelo.xlsx"
    Workbook().save(template)
    existing = tmp_path / "RSAC_032026.xlsx"
    Workbook().save(existing)

    workbook_path = resolve_monthly_workbook(
        template_path=template,
        output_dir=tmp_path,
        competencia="03/2026",
        file_name="RSAC_032026.xlsx",
    )

    assert workbook_path == existing
