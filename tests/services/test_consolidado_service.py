from pathlib import Path

from openpyxl import Workbook, load_workbook

from rsac_relatorios_risco.services.consolidado_service import ConsolidadoService
from rsac_relatorios_risco.services.report_service import ReportData


HEADERS = [
    "Central",
    "Singular",
    "CPF/CNPJ",
    "Nome/Razão Social",
    "Data-base",
    "Classificação de RSAC vigente",
]


def _build_empty_workbook(tmp_path: Path, *, sheet_name: str = "3333") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    path = tmp_path / "consolidado.xlsx"
    workbook.save(path)
    return path


def _sample_report() -> ReportData:
    all_rows = [
        ["SISBR - RISCOS SOCIAL, AMBIENTAL E CLIMÁTICO", None, None, None, None, None],
        ["RELATÓRIO DE RISCO POR COOPERATIVA", None, None, None, None, None],
        ["1004 - SICOOB NOVA CENTRAL", None, None, None, None, None],
        ["DATA DE EMISSÃO", "10/04/2026 11:05", None, None, None, None],
        ["CRITÉRIOS", "DATABASE: 032026, CENTRAL: 1004, SINGULAR: 3333", None, None, None, None],
        HEADERS,
        [
            "1004 - SICOOB NOVA CENTRAL",
            "3333 - SICOOB SECOVICRED",
            "111.111.111-11",
            "Novo Nome 1",
            "03/2026",
            "Baixo",
        ],
        [
            "1004 - SICOOB NOVA CENTRAL",
            "3333 - SICOOB SECOVICRED",
            "222.222.222-22",
            "Novo Nome 2",
            "03/2026",
            "Alto",
        ],
    ]
    return ReportData(
        cooperativa="3333",
        competencia="03/2026",
        headers=HEADERS,
        rows=[all_rows[6], all_rows[7]],
        data_emissao="10/04/2026 11:05",
        criterios="DATABASE: 032026, CENTRAL: 1004, SINGULAR: 3333",
        all_rows=all_rows,
    )


def test_apply_report_copies_all_insumo_content_to_sheet(tmp_path: Path):
    workbook_path = _build_empty_workbook(tmp_path)

    service = ConsolidadoService()
    service.apply_report(workbook_path, _sample_report())

    wb = load_workbook(workbook_path)
    sheet = wb["3333"]

    # Row 1: titulo
    assert sheet["A1"].value == "SISBR - RISCOS SOCIAL, AMBIENTAL E CLIMÁTICO"
    # Row 4: data emissao
    assert sheet["A4"].value == "DATA DE EMISSÃO"
    assert sheet["B4"].value == "10/04/2026 11:05"
    # Row 5: criterios
    assert sheet["B5"].value == "DATABASE: 032026, CENTRAL: 1004, SINGULAR: 3333"
    # Row 6: headers
    assert sheet["A6"].value == "Central"
    assert sheet["F6"].value == "Classificação de RSAC vigente"
    # Row 7-8: dados
    assert sheet["C7"].value == "111.111.111-11"
    assert sheet["D8"].value == "Novo Nome 2"
    assert sheet["F8"].value == "Alto"
    # Row 9: vazio (sem dados alem dos copiados)
    assert sheet["A9"].value is None


def test_apply_report_clears_old_content_before_writing(tmp_path: Path):
    workbook_path = _build_empty_workbook(tmp_path)
    wb = load_workbook(workbook_path)
    sheet = wb["3333"]
    sheet["A1"] = "Dado antigo"
    sheet["A2"] = "Outro dado antigo"
    sheet["Z30"] = "Dado distante"
    wb.save(workbook_path)

    service = ConsolidadoService()
    service.apply_report(workbook_path, _sample_report())

    wb = load_workbook(workbook_path)
    sheet = wb["3333"]

    # Conteudo antigo foi limpo
    assert sheet["Z30"].value is None
    # Novo conteudo presente
    assert sheet["A1"].value == "SISBR - RISCOS SOCIAL, AMBIENTAL E CLIMÁTICO"
